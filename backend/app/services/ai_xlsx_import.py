"""AI-powered XLSX import: sends raw spreadsheet data to Grok for interpretation."""

import io
import json
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.models.exercise import Exercise


def _detect_header_row(rows: list[tuple]) -> int | None:
    """Find the header row index by looking for common column names."""
    keywords = {"exercise", "date", "weight", "rep", "reps", "set", "rpe", "actual", "%"}
    for idx, row in enumerate(rows[:10]):
        cells_lower = {str(c).strip().lower() for c in row if c is not None}
        if len(cells_lower & keywords) >= 2:
            return idx
    return None


def _detect_data_columns(header: tuple) -> list[int]:
    """Return indices of columns that contain training data (not right-side metadata).

    Filters out metadata columns like 'Current Max', 'Future Max', '% Increase'.
    """
    meta_keywords = {"current max", "future max", "% increase", "increase", "max"}
    indices: list[int] = []
    for i, cell in enumerate(header):
        label = str(cell).strip().lower() if cell else ""
        if any(kw in label for kw in meta_keywords):
            continue
        indices.append(i)
    return indices


def _sheet_has_actual_data(rows: list[tuple], header_idx: int | None) -> bool:
    """Check whether a sheet has completed training data (dates or 'Y' actuals)."""
    headers = [str(c).strip().lower() if c else "" for c in rows[header_idx]] if header_idx is not None else []
    date_col = next((i for i, h in enumerate(headers) if h == "date"), None)
    actual_col = next((i for i, h in enumerate(headers) if h == "actual"), None)

    if date_col is None and actual_col is None:
        return False  # No date column and no actual column — likely a template

    # Check if at least some rows have dates or 'Y' in actual
    data_rows = rows[header_idx + 1:] if header_idx is not None else rows[1:]
    has_dates = False
    has_actuals = False
    for row in data_rows[:50]:
        if date_col is not None and date_col < len(row):
            val = str(row[date_col]).strip() if row[date_col] else ""
            if val and val.lower() != "none":
                has_dates = True
        if actual_col is not None and actual_col < len(row):
            val = str(row[actual_col]).strip().upper() if row[actual_col] else ""
            if val == "Y":
                has_actuals = True
        if has_dates or has_actuals:
            return True
    return False


def _fmt(val: Any) -> str:
    """Format a cell value as a clean string."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def extract_sheet_cells(ws: Any, max_rows: int = 500) -> str:
    """Extract raw cell data from a single worksheet as readable text.

    Returns empty string if the sheet appears to be a template with no actual data.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ""

    # Trim trailing empty rows
    while rows and all(c is None or str(c).strip() == "" for c in rows[-1]):
        rows.pop()
    if not rows:
        return ""

    # Detect header and check if this sheet has real data
    header_idx = _detect_header_row(rows)
    if header_idx is not None and not _sheet_has_actual_data(rows, header_idx):
        return ""  # Template-only sheet, skip

    # Determine which columns are training data (not metadata)
    if header_idx is not None:
        data_cols = _detect_data_columns(rows[header_idx])
    else:
        data_cols = list(range(len(rows[0])))

    rows = rows[:max_rows]

    lines: list[str] = [f"=== Sheet: {ws.title} ==="]
    for row_idx, row in enumerate(rows, start=1):
        cells = [_fmt(row[i]) if i < len(row) else "" for i in data_cols]
        if all(c == "" for c in cells):
            continue
        lines.append(f"Row {row_idx}: {' | '.join(cells)}")

    return "\n".join(lines)


async def get_exercise_names(db: AsyncSession) -> list[str]:
    """Get all exercise names from the database for matching guidance."""
    result = await db.execute(select(Exercise.name))
    return [row[0] for row in result.all()]


def build_import_system_prompt(exercise_names: list[str]) -> str:
    """Build the system prompt that instructs Grok to interpret spreadsheet data."""
    exercises_list = "\n".join(f"  - {name}" for name in exercise_names) if exercise_names else "  (no exercises in database yet)"

    return f"""You are a data extraction assistant for a fitness training tracker app.

Your job: Read raw spreadsheet data and convert it into structured workout JSON.

The spreadsheet may have ANY format. Be generous — import everything you can.

== WEIGHT TEXT PARSING ==
Convert text weight values to numeric lbs:
- "BW" or "Bodyweight" → 0
- "Band" or "Bands" → 0
- "25lb DB's", "25lb DB", "25lb DBs", "25 lb DB" → 25 (use single dumbbell weight)
- "15lb DB's" → 15, "30lb DB" → 30, "35lb DB's" → 35, "20lb DB's" → 20
- "5lbs", "8lbs", "15lbs", "35lbs" → extract the number (5, 8, 15, 35)
- "135 lbs", "135lb", "135" → 135
- Any number followed by lb/lbs/pound/pounds → extract the number
- If weight looks like kg (e.g., "100kg") → multiply by 2.20462, add warning

== SETS/REPS TEXT PARSING ==
The "Sets" column often contains both set count and rep count as text:
- "3 sets - 25reps ea"  → 3 sets × 25 reps each
- "3sets - 25reps ea"   → 3 sets × 25 reps each
- "3 sets - 10reps ea"  → 3 sets × 10 reps each
- "3 sets - 15reps ea"  → 3 sets × 15 reps each
- "1set - 2min hold"    → 1 set × 1 rep (time-based, treat reps=1)
- "3sets - 2min hold"   → 3 sets × 1 rep (time-based)
- "3 sets - 2mins ea"   → 3 sets × 1 rep (time-based)
- "3sets - 3sec hold - 10reps ea → ..." → 3 sets × 10 reps (ignore timing)
- "10secs - 10reps ea → 3min10secs/set" → 3 sets × 10 reps
- "3sets - 10secs - 10reps ea → ..." → 3 sets × 10 reps
- "3sets - 3sec hold - 10reps ea → ..." → 3 sets × 10 reps
- "1set - 2min hold"    → 1 set × 1 rep
- "N x M" or "NxM"     → N sets × M reps
- "3 x 5"              → 3 sets × 5 reps

Rule: pattern is always "N sets - Mreps ea". Timed sets (hold, min, secs) without reps → reps=1.

== DATE INFERENCE FROM DAY-LABELED SECTIONS ==
Many spreadsheets organize by week:
1. Look for "Wk of:" or "Week of:" followed by a date in the first 3 rows → this is the week start (Monday)
2. Day labels appear as single-cell rows: Mon, Tues, Tue, Wed, Thurs, Thu, Fri, Sat, Sun
3. Map day to date: Mon=+0, Tue/Tues=+1, Wed=+2, Thu/Thurs=+3, Fri=+4, Sat=+5, Sun=+6
4. All exercises between a day label and the next day label belong to that day's date

== SECTION STRUCTURE ==
A sheet may have multiple sections per day, each with a header row:
- "Run" section: columns are (None/Run, Miles, Time, Avg Pace, Cals, Notes) → create exercise named "Run", weight_lbs=0, reps=1, put distance/time/pace in notes
- "Rehab Workout", "Strength Workout", etc.: standard exercise rows follow
  - Col 0: Exercise name
  - Col 1: Weight (parse as above)
  - Col 2: Sets/reps text (parse as above)
  - Col 3: Rest time (ignore)
  - Col 4-6: Individual set rep counts (may be empty; fall back to Col 2 parsing)
  - Col 7: Notes (include as exercise notes if non-empty)

Skip rows that are:
- Day label rows (single cell: Mon, Tues, etc.)
- Column header rows (containing words like "Weight", "Sets", "Rest", "Reps")
- Section header rows ("Rehab Workout", "Medical Notes:", "Training Notes:", etc.)
- Fully empty rows
- Footer/notes rows at the bottom

== SHORTHAND EXERCISE MATCHING ==
Match shorthand names: "Bench"=bench press, "Squat"=squat, "DL"=deadlift, "OHP"=overhead press,
"Chins"=chin-ups, "RDL"=romanian deadlift, "DB Bench"=dumbbell bench press, etc.

== KNOWN EXERCISES IN THE DATABASE ==
{exercises_list}

Use the EXACT name from the list above when recognized. Otherwise use the name as written.

== OUTPUT FORMAT ==
Return ONLY valid JSON (no markdown, no explanation):
{{
  "workouts": [
    {{
      "date": "YYYY-MM-DD",
      "bodyweight": null,
      "sleep_quality": null,
      "fatigue_level": null,
      "exercises": [
        {{
          "name": "Exercise Name",
          "sets": [
            {{
              "set_number": 1,
              "weight_lbs": 25,
              "reps": 25,
              "rpe": null,
              "set_type": "working",
              "notes": null
            }}
          ]
        }}
      ]
    }}
  ],
  "warnings": ["any issues or assumptions you made"]
}}

== RULES ==
- Create one workout per day. Group all exercises from that day into it.
- Expand sets: "3 sets - 25reps ea" → 3 separate set objects (set_number 1, 2, 3) each with reps=25.
- set_type: "working" by default; "warmup" for warm-up/stretch exercises.
- RPE: null unless explicitly stated. Do NOT infer RPE.
- Be GENEROUS: when in doubt, import the exercise with best-guess values and add a warning.
- "Actual" column with "Y" = completed; include only those rows if present. If no "Actual" column, include all exercise rows.
- Rep scheme "2 x 1" → 2 sets × 1 rep. Always expand into individual set objects.
- Return ONLY the JSON object. No markdown fences."""


async def _call_grok_import(system_prompt: str, user_message: str) -> str:
    """Call Grok with higher token limit for import tasks."""
    import httpx

    if not settings.xai_api_key:
        raise ValueError("XAI_API_KEY not configured")

    async with httpx.AsyncClient(timeout=120.0) as client:
        response = await client.post(
            f"{settings.xai_base_url}/chat/completions",
            headers={
                "Authorization": f"Bearer {settings.xai_api_key}",
                "Content-Type": "application/json",
            },
            json={
                "model": settings.xai_model,
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_message},
                ],
                "temperature": 0.3,
                "max_tokens": 16000,
            },
        )
        response.raise_for_status()
        data = response.json()
        return data["choices"][0]["message"]["content"]


def _parse_grok_json(response_text: str) -> tuple[list[dict], list[str]]:
    """Parse Grok's JSON response, stripping markdown fences if present.

    Returns (workouts, warnings). On parse failure returns ([], [error_message]).
    """
    cleaned = response_text.strip()
    if cleaned.startswith("```"):
        lines = [l for l in cleaned.split("\n") if not l.strip().startswith("```")]
        cleaned = "\n".join(lines)
    try:
        parsed = json.loads(cleaned)
        return parsed.get("workouts", []), parsed.get("warnings", [])
    except json.JSONDecodeError:
        return [], [
            "AI could not parse this sheet into structured data.",
            f"Raw response: {response_text[:300]}",
        ]


async def ai_parse_xlsx(file_bytes: bytes, db: AsyncSession) -> dict[str, Any]:
    """Use Grok to interpret a freeform spreadsheet and return structured workout data.

    Processes each sheet independently and merges results.
    Returns the same format as parse_xlsx() so the existing confirm flow can be reused.
    """
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheets = wb.worksheets
    wb.close()

    # Re-open non-read-only for sheet iteration (read_only doesn't support len well)
    wb2 = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)

    # Get known exercises once for all sheets
    exercise_names = await get_exercise_names(db)
    system_prompt = build_import_system_prompt(exercise_names)

    all_workouts: list[dict] = []
    all_warnings: list[str] = []
    sheets_processed = 0

    for ws in wb2.worksheets:
        sheet_text = extract_sheet_cells(ws)
        if not sheet_text.strip():
            continue

        sheets_processed += 1
        user_message = f"Parse this spreadsheet data into structured workouts:\n\n{sheet_text}"
        response_text = await _call_grok_import(system_prompt, user_message)
        workouts, warnings = _parse_grok_json(response_text)

        if warnings:
            all_warnings.extend([f"[{ws.title}] {w}" for w in warnings])
        all_workouts.extend(workouts)

    wb2.close()

    if sheets_processed == 0:
        return {
            "workouts": [],
            "unmatched_exercises": [],
            "exercise_suggestions": {},
            "warnings": ["Spreadsheet appears to be empty or contains no training data."],
            "stats": {"total_workouts": 0, "total_sets": 0, "date_range": ""},
        }

    workouts = all_workouts
    ai_warnings = all_warnings

    # Load exercises for fuzzy matching
    result = await db.execute(select(Exercise))
    db_exercises: list[Exercise] = list(result.scalars().all())
    exercise_name_map = {ex.name.lower(): ex for ex in db_exercises}

    # Post-process: match exercise names to DB and add order_index
    unmatched: set[str] = set()
    total_sets = 0
    dates: list[str] = []

    for workout in workouts:
        if workout.get("date") and workout["date"] != "unknown":
            dates.append(workout["date"])

        for ex_idx, ex in enumerate(workout.get("exercises", [])):
            ex["order_index"] = ex_idx
            name = ex.get("name", "")
            name_lower = name.strip().lower()

            # Try exact match
            matched_ex = exercise_name_map.get(name_lower)

            # Try contains match
            if not matched_ex:
                for db_ex in db_exercises:
                    db_lower = db_ex.name.lower()
                    if db_lower in name_lower or name_lower in db_lower:
                        matched_ex = db_ex
                        break

            ex["matched_exercise_id"] = matched_ex.id if matched_ex else None
            if not matched_ex:
                unmatched.add(name)

            # Ensure sets have set_number
            for si, s in enumerate(ex.get("sets", []), start=1):
                if not s.get("set_number"):
                    s["set_number"] = si
                total_sets += 1

    # Build stats
    dates.sort()
    date_range = ""
    if dates:
        date_range = f"{dates[0]} to {dates[-1]}" if len(dates) > 1 else dates[0]

    exercise_suggestions: dict[str, int | None] = {name: None for name in sorted(unmatched)}

    return {
        "workouts": workouts,
        "unmatched_exercises": sorted(unmatched),
        "exercise_suggestions": exercise_suggestions,
        "warnings": ai_warnings,
        "stats": {
            "total_workouts": len(workouts),
            "total_sets": total_sets,
            "date_range": date_range,
        },
    }
