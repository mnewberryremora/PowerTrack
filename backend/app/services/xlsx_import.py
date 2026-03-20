"""XLSX import service for parsing training log spreadsheets into structured workout data."""

import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.exercise import Exercise
from app.models.workout import Workout, WorkoutExercise, Set

# ---------------------------------------------------------------------------
# Column name aliases (all compared case-insensitively)
# ---------------------------------------------------------------------------

DATE_ALIASES = {"date", "training date", "workout date", "day"}
EXERCISE_ALIASES = {"exercise", "exercise name", "movement", "lift"}
SET_ALIASES = {"set", "set #", "set number", "set_number"}
WEIGHT_ALIASES = {"weight", "weight (lbs)", "weight_lbs", "load", "load (lbs)", "lbs"}
REPS_ALIASES = {"reps", "repetitions", "rep"}
RPE_ALIASES = {"rpe", "rpe rating", "effort"}
TYPE_ALIASES = {"type", "set type", "set_type"}
NOTES_ALIASES = {"notes", "note", "comments", "comment"}
BODYWEIGHT_ALIASES = {"bodyweight", "body weight", "bw", "bw (lbs)"}
SLEEP_ALIASES = {"sleep", "sleep quality", "sleep_quality"}
FATIGUE_ALIASES = {"fatigue", "fatigue level", "fatigue_level"}
WORKOUT_NAME_ALIASES = {"workout name", "workout", "session", "session name", "name"}

_ALL_ALIAS_GROUPS: list[tuple[str, set[str]]] = [
    ("date", DATE_ALIASES),
    ("exercise", EXERCISE_ALIASES),
    ("set", SET_ALIASES),
    ("weight", WEIGHT_ALIASES),
    ("reps", REPS_ALIASES),
    ("rpe", RPE_ALIASES),
    ("type", TYPE_ALIASES),
    ("notes", NOTES_ALIASES),
    ("bodyweight", BODYWEIGHT_ALIASES),
    ("sleep", SLEEP_ALIASES),
    ("fatigue", FATIGUE_ALIASES),
    ("workout_name", WORKOUT_NAME_ALIASES),
]


def _match_columns(headers: list[str]) -> dict[str, int | None]:
    """Map logical column names to 0-based column indices using alias matching."""
    mapping: dict[str, int | None] = {name: None for name, _ in _ALL_ALIAS_GROUPS}
    for idx, raw_header in enumerate(headers):
        if raw_header is None:
            continue
        normalised = str(raw_header).strip().lower()
        for logical_name, aliases in _ALL_ALIAS_GROUPS:
            if normalised in aliases and mapping[logical_name] is None:
                mapping[logical_name] = idx
                break
    return mapping


def _parse_date(value: Any) -> date | None:
    """Coerce a cell value to a date object."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def _parse_int(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Fuzzy exercise matching
# ---------------------------------------------------------------------------

def _fuzzy_match_exercise(name: str, db_exercises: list[Exercise]) -> Exercise | None:
    """Return the best-matching DB exercise for *name*, or None.

    Strategy:
    1. Exact match (case-insensitive)
    2. Spreadsheet name contains DB name or vice-versa (case-insensitive)
    """
    lower = name.strip().lower()
    # Exact
    for ex in db_exercises:
        if ex.name.lower() == lower:
            return ex
    # Contains
    for ex in db_exercises:
        ex_lower = ex.name.lower()
        if ex_lower in lower or lower in ex_lower:
            return ex
    return None


# ---------------------------------------------------------------------------
# Core parsing
# ---------------------------------------------------------------------------

async def parse_xlsx(file_bytes: bytes, db: AsyncSession) -> dict[str, Any]:
    """Parse an XLSX file and return structured import data.

    Returns::

        {
            "workouts": [
                {
                    "date": "2024-01-15",
                    "bodyweight": 200.0 | None,
                    "sleep_quality": 8 | None,
                    "fatigue_level": 5 | None,
                    "exercises": [
                        {
                            "name": "Competition Squat",
                            "matched_exercise_id": 3 | None,
                            "sets": [
                                {"set_number": 1, "weight_lbs": 315, "reps": 5, "rpe": 7, "set_type": "working", "notes": ""}
                            ]
                        }
                    ]
                }
            ],
            "unmatched_exercises": ["Some Obscure Lift"],
            "exercise_suggestions": {"Some Obscure Lift": null},
            "warnings": [...],
            "stats": {"total_workouts": N, "total_sets": N, "date_range": "..."}
        }
    """
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return {"workouts": [], "warnings": ["Workbook has no active sheet"], "stats": {}}

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"workouts": [], "warnings": ["Sheet has no data rows"], "stats": {}}

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_map = _match_columns(headers)
    warnings: list[str] = []

    # Validate required columns
    if col_map["date"] is None:
        warnings.append("No 'Date' column found — cannot parse file.")
        return {"workouts": [], "warnings": warnings, "stats": {}}
    if col_map["exercise"] is None:
        warnings.append("No 'Exercise' column found — cannot parse file.")
        return {"workouts": [], "warnings": warnings, "stats": {}}
    if col_map["weight"] is None:
        warnings.append("No 'Weight' column found — cannot parse file.")
        return {"workouts": [], "warnings": warnings, "stats": {}}
    if col_map["reps"] is None:
        warnings.append("No 'Reps' column found — cannot parse file.")
        return {"workouts": [], "warnings": warnings, "stats": {}}

    has_set_col = col_map["set"] is not None
    has_type_col = col_map["type"] is not None

    # Load all exercises from DB for fuzzy matching
    result = await db.execute(select(Exercise))
    db_exercises: list[Exercise] = list(result.scalars().all())

    # Parse data rows ----------------------------------------------------------
    # Intermediate: dict keyed by date -> list of row dicts
    date_groups: dict[date, list[dict[str, Any]]] = {}
    # Track per-date metadata (bodyweight, sleep, fatigue) — take first non-null
    date_meta: dict[date, dict[str, Any]] = {}

    total_sets = 0
    for row_idx, row in enumerate(rows[1:], start=2):
        # Skip fully empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        def _cell(logical: str) -> Any:
            idx = col_map[logical]
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        parsed_date = _parse_date(_cell("date"))
        if parsed_date is None:
            warnings.append(f"Row {row_idx}: could not parse date, skipping.")
            continue

        exercise_name = _cell("exercise")
        if exercise_name is None or str(exercise_name).strip() == "":
            warnings.append(f"Row {row_idx}: empty exercise name, skipping.")
            continue
        exercise_name = str(exercise_name).strip()

        weight = _parse_decimal(_cell("weight"))
        reps = _parse_int(_cell("reps"))
        if weight is None or reps is None:
            warnings.append(f"Row {row_idx}: could not parse weight/reps, skipping.")
            continue

        set_number = _parse_int(_cell("set")) if has_set_col else None
        rpe = _parse_decimal(_cell("rpe"))
        set_type = str(_cell("type")).strip().lower() if has_type_col and _cell("type") else "working"
        notes = str(_cell("notes")).strip() if _cell("notes") else None

        workout_name = str(_cell("workout_name")).strip() if _cell("workout_name") else None

        set_dict = {
            "set_number": set_number,  # may be None; auto-numbered later
            "weight_lbs": float(weight),
            "reps": reps,
            "rpe": float(rpe) if rpe is not None else None,
            "set_type": set_type,
            "notes": notes,
            "exercise_name": exercise_name,
            "workout_name": workout_name,
        }

        date_groups.setdefault(parsed_date, []).append(set_dict)
        total_sets += 1

        # Per-date metadata
        if parsed_date not in date_meta:
            date_meta[parsed_date] = {}
        meta = date_meta[parsed_date]
        if "bodyweight" not in meta:
            bw = _parse_decimal(_cell("bodyweight"))
            if bw is not None:
                meta["bodyweight"] = float(bw)
        if "sleep_quality" not in meta:
            sq = _parse_int(_cell("sleep"))
            if sq is not None:
                meta["sleep_quality"] = sq
        if "fatigue_level" not in meta:
            fl = _parse_int(_cell("fatigue"))
            if fl is not None:
                meta["fatigue_level"] = fl

    # Build workouts -----------------------------------------------------------
    exercise_match_cache: dict[str, int | None] = {}
    unmatched_exercises: set[str] = set()

    workouts: list[dict[str, Any]] = []
    for d in sorted(date_groups.keys()):
        rows_for_date = date_groups[d]
        meta = date_meta.get(d, {})
        # Use workout name from first row that has one
        workout_name = next((r["workout_name"] for r in rows_for_date if r.get("workout_name")), None)

        # Group by exercise name, preserving order of first appearance
        exercise_order: list[str] = []
        exercise_sets: dict[str, list[dict]] = {}
        for r in rows_for_date:
            ename = r["exercise_name"]
            if ename not in exercise_sets:
                exercise_order.append(ename)
                exercise_sets[ename] = []
            exercise_sets[ename].append(r)

        exercises_out: list[dict[str, Any]] = []
        for order_idx, ename in enumerate(exercise_order):
            sets = exercise_sets[ename]

            # Auto-number sets if needed
            for i, s in enumerate(sets, start=1):
                if s["set_number"] is None:
                    s["set_number"] = i

            # Fuzzy match exercise
            if ename not in exercise_match_cache:
                matched = _fuzzy_match_exercise(ename, db_exercises)
                exercise_match_cache[ename] = matched.id if matched else None
                if matched is None:
                    unmatched_exercises.add(ename)

            exercises_out.append({
                "name": ename,
                "matched_exercise_id": exercise_match_cache[ename],
                "order_index": order_idx,
                "sets": [
                    {
                        "set_number": s["set_number"],
                        "weight_lbs": s["weight_lbs"],
                        "reps": s["reps"],
                        "rpe": s["rpe"],
                        "set_type": s["set_type"],
                        "notes": s["notes"],
                    }
                    for s in sets
                ],
            })

        workouts.append({
            "date": d.isoformat(),
            "name": workout_name,
            "bodyweight": meta.get("bodyweight"),
            "sleep_quality": meta.get("sleep_quality"),
            "fatigue_level": meta.get("fatigue_level"),
            "exercises": exercises_out,
        })

    # Build exercise suggestions for unmatched (best-effort partial matches)
    exercise_suggestions: dict[str, int | None] = {}
    for name in sorted(unmatched_exercises):
        exercise_suggestions[name] = None  # frontend must resolve

    # Stats
    dates = sorted(date_groups.keys())
    date_range = ""
    if dates:
        date_range = f"{dates[0].isoformat()} to {dates[-1].isoformat()}" if len(dates) > 1 else dates[0].isoformat()

    wb.close()
    return {
        "workouts": workouts,
        "unmatched_exercises": sorted(unmatched_exercises),
        "exercise_suggestions": exercise_suggestions,
        "warnings": warnings,
        "stats": {
            "total_workouts": len(workouts),
            "total_sets": total_sets,
            "date_range": date_range,
        },
    }


# ---------------------------------------------------------------------------
# Create DB records from confirmed import
# ---------------------------------------------------------------------------

async def create_workouts_from_import(
    db: AsyncSession,
    import_data: list[dict[str, Any]],
    exercise_map: dict[str, int],
    user_id: int = 0,
) -> dict[str, Any]:
    """Create Workout + WorkoutExercise + Set records from parsed import data.

    Parameters
    ----------
    import_data:
        The ``workouts`` list returned by ``parse_xlsx`` (or a subset the user confirmed).
    exercise_map:
        User-confirmed mapping of spreadsheet exercise names to DB exercise IDs.
        Every exercise name that appears in *import_data* must have a mapping.

    Returns
    -------
    dict with ``created_workout_ids``, ``created`` (count), and ``errors``.
    """
    created_ids: list[int] = []
    errors: list[str] = []

    for workout_data in import_data:
        try:
            workout_date = date.fromisoformat(workout_data["date"])
            workout = Workout(
                user_id=user_id,
                date=workout_date,
                name=workout_data.get("name") or f"Imported {workout_date.isoformat()}",
                bodyweight_lbs=Decimal(str(workout_data["bodyweight"])) if workout_data.get("bodyweight") else None,
                sleep_quality=workout_data.get("sleep_quality"),
                fatigue_level=workout_data.get("fatigue_level"),
                completed=True,
            )
            db.add(workout)
            await db.flush()

            for ex_data in workout_data.get("exercises", []):
                ex_name = ex_data["name"]
                exercise_id = exercise_map.get(ex_name)
                if exercise_id is None:
                    errors.append(f"No exercise mapping for '{ex_name}' on {workout_data['date']}, skipping exercise.")
                    continue

                we = WorkoutExercise(
                    workout_id=workout.id,
                    exercise_id=exercise_id,
                    order_index=ex_data.get("order_index", 0),
                )
                db.add(we)
                await db.flush()

                for set_data in ex_data.get("sets", []):
                    s = Set(
                        workout_exercise_id=we.id,
                        set_number=set_data["set_number"],
                        weight_lbs=Decimal(str(set_data["weight_lbs"])),
                        reps=set_data["reps"],
                        rpe=Decimal(str(set_data["rpe"])) if set_data.get("rpe") is not None else None,
                        set_type=set_data.get("set_type", "working"),
                        notes=set_data.get("notes"),
                    )
                    db.add(s)

            await db.flush()
            created_ids.append(workout.id)
        except Exception as exc:
            errors.append(f"Error creating workout for {workout_data.get('date', '?')}: {exc}")

    if created_ids:
        await db.commit()

    return {
        "created_workout_ids": created_ids,
        "created": len(created_ids),
        "errors": errors,
    }


# ---------------------------------------------------------------------------
# Template generator
# ---------------------------------------------------------------------------

def generate_template() -> bytes:
    """Generate a template XLSX file with headers and example data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Training Log"

    headers = ["Date", "Exercise", "Set", "Weight (lbs)", "Reps", "RPE", "Type", "Notes"]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Example data
    example_rows = [
        ["2024-01-15", "Competition Squat", 1, 315, 5, 7.0, "working", "Felt good"],
        ["2024-01-15", "Competition Squat", 2, 335, 5, 8.0, "working", ""],
        ["2024-01-15", "Competition Squat", 3, 355, 3, 9.0, "working", "Heavy"],
        ["2024-01-15", "Competition Bench Press", 1, 225, 5, 7.0, "working", ""],
        ["2024-01-15", "Competition Bench Press", 2, 245, 5, 8.0, "working", ""],
        ["2024-01-15", "Competition Bench Press", 3, 255, 3, 8.5, "working", "Paused"],
        ["2024-01-17", "Competition Deadlift", 1, 405, 5, 7.0, "working", ""],
        ["2024-01-17", "Competition Deadlift", 2, 425, 3, 8.5, "working", ""],
        ["2024-01-17", "Competition Deadlift", 3, 455, 1, 9.5, "working", "Grinder"],
    ]
    for row in example_rows:
        ws.append(row)

    # Auto-fit column widths (approximate)
    col_widths = [14, 28, 6, 14, 6, 6, 10, 20]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
