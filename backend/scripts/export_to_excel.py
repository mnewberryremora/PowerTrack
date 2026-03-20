"""Export all workout data to an Excel file for import into the web app.

Saves to /app/training_export.xlsx inside the container.
Run with: docker-compose exec backend python scripts/export_to_excel.py
Then copy out: docker cp <container_id>:/app/training_export.xlsx .
"""

import asyncio
import os
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import text
from sqlalchemy.ext.asyncio import create_async_engine

DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql+asyncpg://training:training_secret@postgres:5432/training_app",
)

OUTPUT_PATH = "/app/training_export.xlsx"


async def main():
    engine = create_async_engine(DATABASE_URL)
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")

    async with engine.connect() as conn:
        # ── Exercises lookup ──
        ex_rows = await conn.execute(text("SELECT id, name FROM exercises ORDER BY name"))
        exercises = {row[0]: row[1] for row in ex_rows.fetchall()}

        # ── Workouts sheet ──
        ws = wb.active
        ws.title = "Workouts"
        headers = ["Date", "Workout Name", "Exercise", "Weight (lbs)", "Set #", "Reps", "RPE", "Set Type", "Notes"]
        ws.append(headers)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        workouts = await conn.execute(
            text("SELECT id, date, name, notes, bodyweight_lbs, sleep_quality, fatigue_level, completed FROM workouts ORDER BY date")
        )
        workouts = workouts.fetchall()

        for w_id, w_date, w_name, w_notes, bw, sleep, fatigue, completed in workouts:
            we_rows = await conn.execute(
                text("SELECT id, exercise_id, order_index, notes FROM workout_exercises WHERE workout_id=:wid ORDER BY order_index"),
                {"wid": w_id},
            )
            for we_id, ex_id, order_idx, we_notes in we_rows.fetchall():
                ex_name = exercises.get(ex_id, f"Unknown ({ex_id})")
                sets = await conn.execute(
                    text("SELECT set_number, weight_lbs, reps, rpe, set_type, notes FROM sets WHERE workout_exercise_id=:weid ORDER BY set_number"),
                    {"weid": we_id},
                )
                for set_num, weight, reps, rpe, set_type, s_notes in sets.fetchall():
                    notes_parts = [n for n in [we_notes, s_notes] if n]
                    ws.append([
                        str(w_date),
                        w_name or "",
                        ex_name,
                        float(weight) if weight is not None else 0.0,
                        set_num,
                        reps or 0,
                        float(rpe) if rpe is not None else None,
                        set_type or "working",
                        "; ".join(notes_parts),
                    ])

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        # ── Body Metrics sheet ──
        bm_ws = wb.create_sheet("Body Metrics")
        bm_headers = ["Date", "Bodyweight (lbs)", "Body Fat %", "Notes"]
        bm_ws.append(bm_headers)
        for col, h in enumerate(bm_headers, 1):
            cell = bm_ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

        bm_rows = await conn.execute(
            text("SELECT date, bodyweight_lbs, body_fat_pct, notes FROM body_metrics ORDER BY date")
        )
        for row in bm_rows.fetchall():
            bm_ws.append([str(row[0]), row[1], row[2], row[3]])

        # ── Meets sheet ──
        meets_ws = wb.create_sheet("Meets")
        m_headers = ["Date", "Name", "Location", "Federation", "Weight Class (kg)", "Status",
                     "Squat Opener (lbs)", "Bench Opener (lbs)", "Deadlift Opener (lbs)", "Notes"]
        meets_ws.append(m_headers)
        for col, h in enumerate(m_headers, 1):
            cell = meets_ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

        meet_rows = await conn.execute(
            text("SELECT date, name, location, federation, weight_class_kg, status, squat_opener_lbs, bench_opener_lbs, deadlift_opener_lbs, notes FROM meets ORDER BY date")
        )
        for row in meet_rows.fetchall():
            meets_ws.append([str(row[0])] + list(row[1:]))

    await engine.dispose()
    wb.save(OUTPUT_PATH)
    print(f"✓ Exported {len(workouts)} workouts to {OUTPUT_PATH}")


asyncio.run(main())
